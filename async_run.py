__version__ = "22.12.13.01"
__author__ = "Oren Brigg"
__author_email__ = "obrigg@cisco.com"
__license__ = "Cisco Sample Code License, Version 1.1 - https://developer.cisco.com/site/license/cisco-sample-code-license/"


import os
import sys
import time
import meraki
import asyncio
import requests
import meraki.aio
from getpass import getpass
from rich import print as pp
from rich.table import Table
from openpyxl import Workbook
from rich.console import Console
from openpyxl.styles import Font, Color


def select_org():
    # Fetch and select the organization
    print("\n\nFetching organizations...\n")
    try:
        organizations = dashboard.organizations.getOrganizations()
    except Exception as e:
        pp(f"[red]An error has occured: \n\n{e}[/red]\n\nExiting...")
        pp("A newly generated API key will require up to 15 minutes to synchronize with Meraki API gateways. \
            \nIf you're using a new key - kindly try again in a few minutes.")
        sys.exit(1)
    organizations.sort(key=lambda x: x["name"])
    ids = []
    table = Table(title="Meraki Organizations")
    table.add_column("Organization #", justify="left", style="cyan", no_wrap=True)
    table.add_column("Org Name", justify="left", style="cyan", no_wrap=True)
    counter = 0
    for organization in organizations:
        ids.append(organization["id"])
        table.add_row(str(counter), organization["name"])
        counter += 1
    console = Console()
    console.print(table)
    isOrgDone = False
    while isOrgDone == False:
        selected = input(
            "\nKindly select the organization ID you would like to query: "
        )
        try:
            if int(selected) in range(0, counter):
                isOrgDone = True
            else:
                print("\t[red]Invalid Organization Number\n")
        except:
            print("\t[red]Invalid Organization Number\n")
    return (organizations[int(selected)]["id"], organizations[int(selected)]["name"])


async def async_check_network_health_alerts(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the network health alerts for a given network.
    """
    # print(f"\t\tChecking network health alerts for network: {network['name']}")
    try:
        alerts = await aiomeraki.networks.getNetworkHealthAlerts(network["id"])
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    if len(alerts) == 0:
        pp(f"[green]No network health alerts for network {network['name']}")
        results[network["name"]]["network_health_alerts"] = {"is_ok": True}
    else:
        result = {"is_ok": False, "alert_list": []}
        pp(f"[red]Network alerts detected for network {network['name']}")
        for alert in alerts:
            try:
                del alert["scope"]["devices"][0]["url"]
                del alert["scope"]["devices"][0]["mac"]
            except:
                pass
            result["alert_list"].append(
                {
                    "severity": alert["severity"],
                    "category": alert["category"],
                    "type": alert["type"],
                    "details": alert["scope"],
                }
            )
            pp(
                f"[red]Severity: {alert['severity']}\tCategory: {alert['category']}\tType: {alert['type']}"
            )
        results[network["name"]]["network_health_alerts"] = result


async def async_check_wifi_channel_utilization(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the wifi channel utilization for a given network.
    if the channel utilization is above the threshold, the check will fail.

    it will populate "results" with a dictionary with the result for each AP.
    e.g. {
    'is_ok': False,
    'Q2KD-XXXX-XXXX': {'is_ok': False, 'name': 'AP1', 'utilization': 51.66, 'occurances': 3},
    'Q2KD-XXXX-XXXX': {'is_ok': False, 'name': 'AP2', 'utilization': 56.69, 'occurances': 17},
    'Q2KD-XXXX-XXXX': {'is_ok': True, 'name': 'AP3', 'utilization': 16.93, 'occurances': 8},
    'Q2KD-XXXX-XXXX': {'is_ok': False, 'name': 'AP4', 'utilization': 59.48, 'occurances': 1}
    }
    """
    # print(f"\t\tChecking wifi channel utilization for network: {network['name']}")
    result = {"is_ok": True}
    try:
        channel_utilization = (
            await aiomeraki.networks.getNetworkNetworkHealthChannelUtilization(
                network["id"], perPage=100
            )
        )
        # TODO: pagination
        for ap in channel_utilization:
            utilization_list = [
                ap["wifi1"][util]["utilization"] for util in range(len(ap["wifi1"]))
            ]
            exceeded_utilization_list = [
                utilization
                for utilization in utilization_list
                if utilization > thresholds["5G Channel Utilization"]
            ]
            if len(utilization_list) == 0:
                pp(f"[yellow]AP {ap['serial']} does not have 5GHz enabled. Skipping...")
            elif len(exceeded_utilization_list) > 0:
                pp(
                    f"[red]5GHz Channel Utilization exceeded {thresholds['5G Channel Utilization']}% {len(exceeded_utilization_list)} times, with a peak of {max(utilization_list)}% for AP {ap['serial']}"
                )
                result[ap["serial"]] = {
                    "is_ok": False,
                    "utilization": max(utilization_list),
                    "occurances": len(exceeded_utilization_list),
                }
                result["is_ok"] = False
            else:
                pp(
                    f"[green]5GHz Channel did not exceed {thresholds['5G Channel Utilization']}% for AP {ap['serial']}, max utilization was {max(utilization_list)}"
                )
                result[ap["serial"]] = {
                    "is_ok": True,
                    "utilization": max(utilization_list),
                    "occurances": 0,
                }
        # Adding AP names
        try:
            network_devices = await aiomeraki.networks.getNetworkDevices(network["id"])
        except meraki.exceptions.AsyncAPIError as e:
            pp(
                f'[bold magenta]Meraki AIO API Error (Network Name "{network["name"]}": \n { e }'
            )
        except Exception as e:
            pp(f"[bold magenta]Some other ERROR: {e}")
        for device in network_devices:
            if device["serial"] in result:
                result[device["serial"]]["name"] = device.get("name", device["serial"])
        #
        results[network["name"]]["channel_utilization_check"] = result
    #
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
        results[network["name"]]["channel_utilization_check"] = {"is_ok": False}
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
        results[network["name"]]["channel_utilization_check"] = {"is_ok": False}


async def async_check_wifi_rf_profiles(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the RF profiles for a given network.

    it will populate "results" with a dictionary with the result for each AP.
    e.g.
    {'is_ok': False,
    'RF Profile 1': {
                'is_ok': False,
                'tests': {
                    'min_power': {'is_ok': True, 'value': 8},
                    'min_bitrate': {'is_ok': True, 'value': 54},
                    'channel_width': {'is_ok': True, 'value': '40'},
                    'rxsop': {'is_ok': False, 'value': -75}
                }},
    'RF Profile 2': {
                'is_ok': True,
                'tests': {
                    'min_power': {'is_ok': True, 'value': 5},
                    'min_bitrate': {'is_ok': True, 'value': 54},
                    'channel_width': {'is_ok': False, 'value': 'auto'},
                    'rxsop': {'is_ok': True, 'value': None}
                }}}

    """
    # print(f"\t\tChecking WiFi RF Profiles for network: {network['name']}")
    result = {"is_ok": True}
    try:
        rf_profiles = await aiomeraki.wireless.getNetworkWirelessRfProfiles(
            network["id"]
        )
    except:
        pp(f"[red]Could not fetch RF profiles for network {network['name']}")
        return {
            "is_ok": False,
            "ERROR": {
                "is_ok": False,
                "tests": {
                    "min_power": {"is_ok": False, "value": ""},
                    "min_bitrate": {"is_ok": True, "value": ""},
                    "channel_width": {"is_ok": True, "value": ""},
                    "rxsop": {"is_ok": False, "value": ""},
                },
            },
        }
    for rf_profile in rf_profiles:
        result[rf_profile["name"]] = {
            "is_ok": True,
            "tests": {
                "min_power": {"is_ok": True},
                "min_bitrate": {"is_ok": True},
                "channel_width": {"is_ok": True},
                "rxsop": {"is_ok": True},
            },
        }
        # Check min TX power
        if rf_profile["fiveGhzSettings"]["minPower"] > thresholds["5G Min TX Power"]:
            pp(
                f"[red]The min TX power is too high at {rf_profile['fiveGhzSettings']['minPower']}dBm (not including antenna gain) for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["min_power"] = {
                "is_ok": False,
                "value": rf_profile["fiveGhzSettings"]["minPower"],
            }
            result[rf_profile["name"]]["is_ok"] = False
            result["is_ok"] = False
        else:
            pp(
                f"[green]The min TX power is {rf_profile['fiveGhzSettings']['minPower']}dBm for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["min_power"] = {
                "is_ok": True,
                "value": rf_profile["fiveGhzSettings"]["minPower"],
            }

        # Check min bitrate
        if rf_profile["fiveGhzSettings"]["minBitrate"] < thresholds["5G Min Bitrate"]:
            pp(
                f"[red]The min bitrate is {rf_profile['fiveGhzSettings']['minBitrate']}Mbps for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["min_bitrate"] = {
                "is_ok": False,
                "value": rf_profile["fiveGhzSettings"]["minBitrate"],
            }
            result[rf_profile["name"]]["is_ok"] = False
            result["is_ok"] = False
        else:
            pp(
                f"[green]The min bitrate is {rf_profile['fiveGhzSettings']['minBitrate']}Mbps for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["min_bitrate"] = {
                "is_ok": True,
                "value": rf_profile["fiveGhzSettings"]["minBitrate"],
            }

        # Check channel width
        if rf_profile["fiveGhzSettings"]["channelWidth"] == "auto":
            pp(
                f"[red]The channel width is {rf_profile['fiveGhzSettings']['channelWidth']} for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["channel_width"] = {
                "is_ok": False,
                "value": rf_profile["fiveGhzSettings"]["channelWidth"],
            }
            result[rf_profile["name"]]["is_ok"] = False
            result["is_ok"] = False
        elif (
            int(rf_profile["fiveGhzSettings"]["channelWidth"])
            > thresholds["5G Max Channel Width"]
        ):
            pp(
                f"[red]The channel width is {rf_profile['fiveGhzSettings']['channelWidth']}MHz for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["channel_width"] = {
                "is_ok": False,
                "value": rf_profile["fiveGhzSettings"]["channelWidth"],
            }
            result[rf_profile["name"]]["is_ok"] = False
            result["is_ok"] = False
        else:
            pp(
                f"[green]The channel width is {rf_profile['fiveGhzSettings']['channelWidth']}MHz for RF profile {rf_profile['name']}"
            )
            result[rf_profile["name"]]["tests"]["channel_width"] = {
                "is_ok": True,
                "value": rf_profile["fiveGhzSettings"]["channelWidth"],
            }

        # Check if rx-sop is confiugred
        if rf_profile["fiveGhzSettings"]["rxsop"] != None:
            pp(f"[red]RX-SOP is configured for RF profile {rf_profile['name']}")
            result[rf_profile["name"]]["tests"]["rxsop"] = {
                "is_ok": False,
                "value": rf_profile["fiveGhzSettings"]["rxsop"],
            }
            result[rf_profile["name"]]["is_ok"] = False
            result["is_ok"] = False
        else:
            pp(f"[green]RX-SOP is not configured for RF profile {rf_profile['name']}")
            result[rf_profile["name"]]["tests"]["rxsop"] = {
                "is_ok": True,
                "value": rf_profile["fiveGhzSettings"]["rxsop"],
            }
    results[network["name"]]["rf_profiles_check"] = result


async def async_check_wifi_ssid_amount(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the amount of SSIDs for a given network.

    e.g. {
    'is_ok': False,
    'amount': 5
    }
    """
    # print("\n\t\tChecking WiFi SSID Amount...\n")
    result = {"is_ok": True}
    try:
        ssid_list = await aiomeraki.wireless.getNetworkWirelessSsids(network["id"])
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    enabled_ssid_counter = 0
    for ssid in ssid_list:
        if ssid["enabled"]:
            enabled_ssid_counter += 1
    result["ssid_amount"] = enabled_ssid_counter
    if enabled_ssid_counter <= thresholds["ssid_amount"]:
        pp(
            f"[green]There are {enabled_ssid_counter} SSIDs enabled for network {network['name']}"
        )
    else:
        pp(
            f"[red]There are {enabled_ssid_counter} SSIDs enabled for network {network['name']}"
        )
        result["is_ok"] = False
    results[network["name"]]["ssid_amount_check"] = result


async def async_check_switches_port_counters(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the port counters for all switches in a given network.

    it will return a dictionary with the result for each switch.
    e.g. {
    'is_ok': False,
    'Switch 1': {'is_ok': False, 'crc': ['3'], 'collision': [], 'broadcast': ['17', '18', '19', '20', '27'], 'multicast': [], 'topology_changes': []},
    'Switch 2': {'is_ok': False, 'crc': [], 'collision': ['5'], 'broadcast': ['1', '14', '49'], 'multicast': [], 'topology_changes': []},
    'Switch 3': {'is_ok': True, 'crc': [], 'collision': [], 'broadcast': [], 'multicast': [], 'topology_changes': []},
    }
    """
    # print(f"\t\tChecking Switch Port Counters for network {network['name']}...")
    try:
        device_list = await aiomeraki.networks.getNetworkDevices(network["id"])
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    results[network["name"]]["port_counters_check"] = {"is_ok": True}
    port_check_task = []
    for device in device_list:
        if "MS" in device["model"] or "C9" in device["model"]:
            if "name" not in device.keys():
                device["name"] = device["serial"]
            port_check_task.append(
                async_check_switch_port_counters(aiomeraki, network, device)
            )
            results[network["name"]]["port_counters_check"][device["name"]] = {
                "is_ok": True
            }
    #
    for task in asyncio.as_completed(port_check_task):
        await task


async def async_check_switch_port_counters(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict, device: dict
):
    # print(f"\t\tChecking Switch Port Counters for network {network['name']}...")
    result = {
        "is_ok": True,
        "crc": [],
        "collision": [],
        "broadcast": [],
        "multicast": [],
        "topology_changes": [],
    }
    try:
        switch_counters = await aiomeraki.switch.getDeviceSwitchPortsStatusesPackets(
            device["serial"]
        )
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    for port in switch_counters:
        for port_counter in port["packets"]:
            # CRC and collision errors
            if port_counter["desc"] == "CRC align errors" and port_counter["total"] > 0:
                pp(
                    f"[red]{port_counter['total']} CRC errors on switch {device['name']} - port {port['portId']}"
                )
                result["crc"].append(port["portId"])
                result["is_ok"] = False
            if port_counter["desc"] == "Collisions" and port_counter["total"] > 0:
                pp(
                    f"[red]{port_counter['total']} collisions on switch {device['name']} - port {port['portId']}"
                )
                result["collision"].append(port["portId"])
                result["is_ok"] = False
            # Broadcast and Multicast rates
            if (
                port_counter["desc"] == "Broadcast"
                and port_counter["ratePerSec"]["total"] > thresholds["broadcast_rate"]
            ):
                pp(
                    f"[red]{port_counter['ratePerSec']['total']} broadcast/s on switch {device['name']} - port {port['portId']}"
                )
                result["broadcast"].append(port["portId"])
                result["is_ok"] = False
            if (
                port_counter["desc"] == "Multicast"
                and port_counter["ratePerSec"]["total"] > thresholds["multicast_rate"]
            ):
                pp(
                    f"[red]{port_counter['ratePerSec']['total']} multicast/s on switch {device['name']} - port {port['portId']}"
                )
                result["multicast"].append(port["portId"])
                result["is_ok"] = False
            # Topology changes
            if (
                port_counter["desc"] == "Topology changes"
                and port_counter["total"] > thresholds["topology_changes"]
            ):
                pp(
                    f"[red]{port_counter['total']} topology changes on switch {device['name']} - port {port['portId']}"
                )
                result["topology_changes"].append(port["portId"])
                result["is_ok"] = False
    if results[network["name"]]["port_counters_check"][device["name"]]["is_ok"]:
        pp(f"[green]No port errors on switch {device['name']}")
    results[network["name"]]["port_counters_check"][device["name"]] = result


async def async_check_switch_stp(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the STP status for a given network.
    """
    # print("\n\t\tChecking Switch STP Status...\n")
    try:
        stp_status = await aiomeraki.switch.getNetworkSwitchStp(network["id"])
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    if stp_status["rstpEnabled"]:
        pp(f"[green]STP is enabled for network {network['name']}")
        results[network["name"]]["stp_check"] = {"is_ok": True}
    else:
        pp(f"[red]STP is disabled for network {network['name']}")
        results[network["name"]]["stp_check"] = {"is_ok": False}


async def async_check_switch_mtu(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the MTU of a given network.
    """
    # print(f"\t\tChecking Switch MTU for network {network['name']}...")
    try:
        mtu = await aiomeraki.switch.getNetworkSwitchMtu(network["id"])
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    if mtu["defaultMtuSize"] == 9578 and mtu["overrides"] == []:
        pp(
            f"[green]Jumbo Frames enabled for network {network['name']} (MTU: {mtu['defaultMtuSize']})"
        )
        results[network["name"]]["mtu_check"] = {"is_ok": True}
    else:
        pp(
            f"[red]Jumbo Frames disabled for network {network['name']} (MTU: {mtu['defaultMtuSize']}).\
            \nIt's recommended to keep at default of 9578 unless intermediate devices don’t support jumbo frames"
        )
        results[network["name"]]["mtu_check"] = {"is_ok": False}


async def async_check_switch_storm_control(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the storm control settings of a given network.
    """
    # print(f"\t\tChecking Switch Storm Control for network: {network['name']}...")
    try:
        storm_control = await aiomeraki.switch.getNetworkSwitchStormControl(
            network["id"]
        )
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    try:
        if (
            storm_control["broadcastThreshold"] < 100
            and storm_control["multicastThreshold"] < 100
            and storm_control["unknownUnicastThreshold"] < 100
        ):
            pp(f"[green]Storm-control is enabled for network {network['name']}.")
            results[network["name"]]["storm_control"] = {"is_ok": True}
        else:
            pp(
                f"[yellow]Storm-control is disabled for network {network['name']}. Best practices suggest a limit should be configured."
            )
            results[network["name"]]["storm_control"] = {"is_ok": False}
    except:
        pp(f"[yellow]Storm-control is not supported for network {network['name']}.")
        results[network["name"]]["storm_control"] = {"is_ok": False}


async def async_check_network_firmware(
    aiomeraki: meraki.aio.AsyncDashboardAPI, network: dict
):
    """
    This fuction checks the firmware versions of a given network.
    e.g. {
        'appliance': {'current_version_name': 'MX 16.15', 'latest_stable_version': 'MX 16.15'},
        'wireless': {'current_version_name': 'MR 27.7.1', 'latest_stable_version': 'MR 28.5'}
    }
    """
    result = {"is_ok": True}
    # print(f"\t\tChecking Firmware Version for network {network['name']}...")
    try:
        response = await aiomeraki.networks.getNetworkFirmwareUpgrades(network["id"])
        firmware = response["products"]
        for product in firmware:
            current_version = firmware[product]["currentVersion"]["shortName"]
            # Looking for the latest stable version
            for version in firmware[product]["availableVersions"]:
                if version["releaseType"] == "stable":
                    latest_stable_version = version["shortName"]
            if current_version == latest_stable_version:
                pp(
                    f"[green]{network['name']}: {product.upper()} is running the current stable version ({current_version})."
                )
            elif firmware[product]["nextUpgrade"]["time"] != "":
                pp(
                    f"[green]{network['name']}: {product.upper()} is not running the current stable version ({current_version}), but an upgrade is scheduled for {firmware[product]['nextUpgrade']['time']}"
                )
            else:
                pp(
                    f"[red]{network['name']}: {product.upper()} is not running the current stable version (current: {current_version}, current stable version: {latest_stable_version})"
                )
                result["is_ok"] = False
            #
            result[product] = {
                "current_version": current_version,
                "latest_stable_version": latest_stable_version,
                "scheduled_upgrade": firmware[product]["nextUpgrade"]["time"],
            }
        results[network["name"]]["network_firmware_check"] = result
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )

    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")


async def async_check_org_admins(aiomeraki: meraki.aio.AsyncDashboardAPI):
    """
    This fuction checks the administration settings of the organization.

    it will return a dictionary with the results for the admin checks.
    e.g. {
    'is_ok': False,
    'more_than_one_admin': True,
    'users': {
        '123456': {'email': 'user1@org.com', 'name': 'user1', '2fa': True, 'api_calls': 50},
        '654321': {'email': 'user2@org.com', 'name': 'user2', '2fa': False, 'api_calls': 50}},
    'missing_2fa': True,
    'api_calls': 127
    }

    """
    # print(f"\t\tAnalyzing organization admins...")
    results["org_settings"] = {
        "is_ok": False,
        "more_than_one_admin": False,
        "users": {},
        "missing_2fa": True,
        "api_calls": 0,
        "using_v0": False,
    }
    try:
        org_admins = await aiomeraki.organizations.getOrganizationAdmins(org_id)
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    for admin in org_admins:
        results["org_settings"]["users"][admin["id"]] = {
            "email": admin["email"],
            "name": admin["name"],
            "2fa": admin["twoFactorAuthEnabled"],
            "api_calls": 0,
            "using_v0": False,
        }
        if admin["twoFactorAuthEnabled"] == False:
            pp(
                f"[yellow]Missing 2FA for admin {admin['name']} (email: {admin['email']})"
            )
        else:
            pp(
                f"[green]Admin {admin['name']} (email: {admin['email']}) has 2FA enabled"
            )
    # Filter full right admins (not just read-only or network specific admins)
    full_right_admins = [admin for admin in org_admins if admin["orgAccess"] == "full"]
    if len(full_right_admins) > 1:
        results["org_settings"]["more_than_one_admin"] = True
        pp(f"[green]More than one admin has full rights. This is recommended.")
    else:
        pp(
            f"[red]Only one admin has full rights. It's recommended to have at least one admin with full rights."
        )
    if (
        results["org_settings"]["more_than_one_admin"] == True
        and results["org_settings"]["missing_2fa"] == []
    ):
        results["org_settings"]["is_ok"] = True
    else:
        results["org_settings"]["is_ok"] = False
    # Check API access
    check_api_calls_tasks = [
        async_check_api_calls(aiomeraki, admin["id"]) for admin in org_admins
    ]
    for task in asyncio.as_completed(check_api_calls_tasks):
        await task
    api_call_count = 0
    for admin in results["org_settings"]["users"]:
        api_call_count += results["org_settings"]["users"][admin]["api_calls"]
    results["org_settings"]["api_calls"] = api_call_count
    pp(
        f"API access usage: {results['org_settings']['api_calls']} API calls during the last week."
    )


async def async_check_api_calls(
    aiomeraki: meraki.aio.AsyncDashboardAPI, org_admin: str
):
    try:
        api_requests = await aiomeraki.organizations.getOrganizationApiRequests(
            org_id, adminId=org_admin, timespan=7 * 86400, perPage=1000, total_pages=10
        )
    except meraki.exceptions.AsyncAPIError as e:
        pp(
            f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
        )
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    for request in api_requests:
        results["org_settings"]["users"][request["adminId"]]["api_calls"] += 1
        if "/v0/" in request["path"]:
            results["org_settings"]["using_v0"] = True
            if not results["org_settings"]["users"][request["adminId"]]["using_v0"]:
                pp(
                    f"[red]Admin {results['org_settings']['users'][request['adminId']]['name']} (email: {results['org_settings']['users'][request['adminId']]['email']}) is using the v0 API"
                )
            results["org_settings"]["users"][request["adminId"]]["using_v0"] = True
            results["org_settings"]["is_ok"] = False


def check_wireless_ports(headers):
    url = f"https://api.meraki.com/api/v1/organizations/{org_id}/wireless/devices/ethernet/statuses"
    try:
        ap_uplinks = requests.get(url, headers=headers).json()
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    try:
        networks = dashboard.organizations.getOrganizationNetworks(org_id)
    except Exception as e:
        pp(f"[bold magenta]Some other ERROR: {e}")
    #
    for ap in ap_uplinks:
        # Translating network ID to network name
        network_id = ap['network']['id']
        network_name = "N/A"
        for network in networks:
            if network['id'] == network_id:
                network_name = network['name']
                break
        # Checking if the AP has 5GHz history, otherwise it won't appear in the results
        if ap["serial"] in results[network_name]["channel_utilization_check"].keys():
            # Checking the APs uplink
            # TODO: Adjust the code for APs with more than one port
            ap_port = ap['ports'][0]
            results[network_name]["channel_utilization_check"][ap["serial"]]["speed"] = ap_port['linkNegotiation']['speed']
            results[network_name]["channel_utilization_check"][ap["serial"]]["duplex"] = ap_port['linkNegotiation']['duplex']
            if ap_port['linkNegotiation']['speed'] == None or ap_port['linkNegotiation']['speed'] < 1000:
                results[network_name]["channel_utilization_check"][ap["serial"]]["is_ok"] = False
            if ap_port['linkNegotiation']['duplex'] == 'half' or ap_port['linkNegotiation']['duplex'] == None:
                results[network_name]["channel_utilization_check"][ap["serial"]]["is_ok"] = False
        

def generate_excel_report(results: dict) -> None:
    print("\n\t\tGenerating an Excel Report...\n")
    ABC = [None, "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    workbook = Workbook()
    sheet = workbook.active
    #
    # Intro tab
    sheet.title = "Introduction"
    sheet["B3"] = "Introduction to the Meraki health check report"
    sheet[
        "B5"
    ] = "Cisco Meraki is an amazing cloud-managed IT solution, simplifying network, security, security cameras, and IoT infrastructure."
    sheet[
        "B6"
    ] = "However, even the most intelligent AI/ML-driven solution is still vulnerable to users misconfiguring various options (usually without reading the documentation)."
    sheet[
        "B7"
    ] = "Misconfiguration can result in an outage, or poor user experience (if you will limit user's traffic to 1Mbps - things will work slowly.. AI won't help there as it's the admin's 'intent')."
    sheet[
        "B9"
    ] = "This report is presenting the alignment between your Meraki networks' state and configuration with Meraki best practices, "
    sheet[
        "B10"
    ] = "and a set of thresholds I have selected based on my personal experience."
    sheet["B12"] = "In the report you will find the following tabs:"
    sheet[
        "B13"
    ] = "1. Summary - This tab presents a summary of the results of the health check."
    sheet[
        "B14"
    ] = "2. Network Health Alerts - This tab presents the dashboard alerts from all networks in a single view."
    sheet[
        "B15"
    ] = f"3. Network Health - This tab presents the Channel Utilization for every wireless AP. We will examine only the 5GHz spectrum; If you are using the 2.4GHz spectrum - it's beyond saving..."
    sheet[
        "C16"
    ] = f"The threshold is set to {thresholds['5G Channel Utilization']}%. APs with utilization above this threshold for many occurrences (10+) may be experiencing RF issues."
    sheet[
        "B17"
    ] = "4. Firmware Upgrade - This tab presents the firmware status for every network. Highlighting networks that require a firmware upgrade."
    sheet[
        "B18"
    ] = "5. RF profiles - This tab presents the (non-default) RF profiles for every network."
    sheet[
        "C19"
    ] = f"Minimum Tx power: Setting the minimum Tx power too high, might result in wireless APs interfering with each other, as they are not allowed to decrease their power. The threshold is set to {thresholds['5G Min TX Power']} dBm."
    sheet[
        "C20"
    ] = f"Minimum bitrate: Broadcasts and Multicasts will be sent over the wireless at this speed. The lower the speed - the more airtime is wasted. The threshold is set to {thresholds['5G Min Bitrate']} Mbps."
    sheet[
        "C21"
    ] = f"Channel Width: Depending on local regulation and wireless AP density, there is a limited number of channels that can be used. In most deployments, channel width of more than {thresholds['5G Max Channel Width']}MHz might cause interference between the wireless APs."
    sheet[
        "C22"
    ] = f"RX-SOP: This is a fine-tuning network design tool that should be used only after consulting an independent wireless expert or Meraki Support. If it's configured - there should be a good reason for it. More details at: https://documentation.meraki.com/MR/Radio_Settings/Receive_Start_of_Packet_(RX-SOP)"
    sheet[
        "B23"
    ] = f"6. Switch port counters: This tab presents every switch in every network."
    sheet[
        "C24"
    ] = f"Ports with CRC errors: We do not expect to see any CRC errors on our network, ports with more than 0 CRC errors will appear here."
    sheet[
        "C25"
    ] = f"Ports with collisions: It's 2022.. we shouldn't be seeing hubs or collisions on our network. Ports with more than 0 collisions will appear here."
    sheet[
        "C26"
    ] = f"Multicasts exceeding threshold: Multicast traffic may be legitimate, we're highlighting ports with more than {thresholds['multicast_rate']} multicasts per second for visibility (and making sure they are legitimate)."
    sheet[
        "C27"
    ] = f"Broadcasts exceeding threshold: Broadcasts above a certain threshold should be looked at, we're highlighting ports with more than {thresholds['broadcast_rate']} broadcasts per second for visibility (and making sure they are legitimate)."
    sheet[
        "C28"
    ] = f"Topology changes exceeding threshold: TCN means something has changed in the STP topology. We're highlighting ports with more than {thresholds['topology_changes']} topology changes for visibility (and making sure they are legitimate)."
    sheet[
        "B29"
    ] = f"7. Organization Settings - This tab presents the organization settings."
    sheet[
        "C30"
    ] = f"Multiple admins: We're looking for a single admin with full rights. If you see more than one admin with full rights - it's recommended to have at least one admin with full rights."
    sheet[
        "C31"
    ] = f"2FA: Two Factor Authentication is an important security mechanism, highly recommended for securing your admin accounts."
    sheet[
        "C32"
    ] = f"API access: presenting which admin users are using the Dashboard API and whether they are using the v0 API which is being deprecated."
    #
    # Summary tab
    workbook.create_sheet("Summary")
    sheet = workbook["Summary"]
    sheet["A1"] = "Organization Name"
    sheet["B1"] = "Network Name"
    sheet["C1"] = "Test Name"
    sheet["D1"] = "Test Result"
    #
    sheet["A2"] = org_name
    sheet["B2"] = "N/A"
    sheet["C2"] = "Organization Settings"
    if results["org_settings"]["is_ok"] == True:
        sheet["D2"] = "OK"
    else:
        sheet["D2"] = "Fail"
        sheet["D2"].font = Font(bold=True, color="00FF0000")
    #
    line = 3
    for network in results:
        if network == "org_settings":
            continue
        for test_name in results[network]:
            sheet[f"A{line}"] = org_name
            sheet[f"B{line}"] = network
            sheet[f"C{line}"] = test_name
            if results[network][test_name]["is_ok"]:
                sheet[f"D{line}"] = "Pass"
            else:
                sheet[f"D{line}"] = "Fail"
                sheet[f"D{line}"].font = Font(bold=True, color="00FF0000")
            line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A1:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Test Result"])
    #
    # Organization Admin tab
    workbook.create_sheet("Organization Admin")
    sheet = workbook["Organization Admin"]
    sheet["A1"] = "Organization Name"
    sheet["A2"] = org_name
    sheet["B1"] = "2+ admins"
    if results["org_settings"]["more_than_one_admin"]:
        sheet["B2"] = "Yes"
    else:
        sheet["B2"] = "No"
        sheet["B2"].font = Font(bold=True, color="00FF0000")
    sheet["C1"] = "Admins missing 2FA"
    sheet["C2"] = (
        str(results["org_settings"]["missing_2fa"])
        if results["org_settings"]["missing_2fa"] != []
        else ""
    )
    sheet["C2"].font = Font(bold=True, color="00FF0000")
    sheet["D1"] = "API Calls (last 7 days)"
    sheet["D2"] = results["org_settings"]["api_calls"]
    sheet["E1"] = "Using API v0 ?"
    if results["org_settings"]["using_v0"]:
        sheet["E2"] = "Yes"
        sheet["E2"].font = Font(bold=True, color="00FF0000")
    else:
        sheet["E2"] = "No"
    #
    sheet["A5"] = "Organization Name"
    sheet["B5"] = "Admin Name"
    sheet["C5"] = "Admin Email"
    sheet["D5"] = "2FA enablement"
    sheet["E5"] = "API Calls (last 7 days)"
    sheet["F5"] = "Using API v0"
    line = 6
    for admin in results["org_settings"]["users"]:
        sheet[f"A{line}"] = org_name
        sheet[f"B{line}"] = results["org_settings"]["users"][admin]["name"]
        sheet[f"C{line}"] = results["org_settings"]["users"][admin]["email"]
        if results["org_settings"]["users"][admin]["2fa"]:
            sheet[f"D{line}"] = "Yes"
        else:
            sheet[f"D{line}"] = "No"
            sheet[f"D{line}"].font = Font(bold=True, color="00FF0000")
        sheet[f"E{line}"] = results["org_settings"]["users"][admin]["api_calls"]
        if results["org_settings"]["users"][admin]["using_v0"]:
            sheet[f"F{line}"] = "Yes"
            sheet[f"F{line}"].font = Font(bold=True, color="00FF0000")
        else:
            sheet[f"F{line}"] = "No"
        line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A5:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Admin Name"])
    #
    # Network Health Alerts tab
    workbook.create_sheet("Network Health Alerts")
    sheet = workbook["Network Health Alerts"]
    sheet["A1"] = "Organization Name"
    sheet["B1"] = "Network Name"
    sheet["C1"] = "Severity"
    sheet["D1"] = "Category"
    sheet["E1"] = "Type"
    sheet["F1"] = "Details"
    line = 2
    #
    for network in results:
        if network == "org_settings":
            continue
        if results[network]["network_health_alerts"]["is_ok"] == True:
            pass
        else:
            for alert in results[network]["network_health_alerts"]["alert_list"]:
                sheet[f"A{line}"] = org_name
                sheet[f"B{line}"] = network
                sheet[f"C{line}"] = alert["severity"]
                sheet[f"D{line}"] = alert["category"]
                sheet[f"E{line}"] = alert["type"]
                sheet[f"F{line}"] = str(alert["details"])
                if alert["severity"] == "critical":
                    for cell in sheet[line:line]:
                        cell.font = Font(bold=True, color="00FF0000")
                elif alert["severity"] == "warning":
                    for cell in sheet[line:line]:
                        cell.font = Font(bold=True, color="00FF9900")
                line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A1:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Network Name"])
    #
    # Network Firmware tab
    workbook.create_sheet("Network Firmware")
    sheet = workbook["Network Firmware"]
    sheet["A1"] = "Organization Name"
    sheet["B1"] = "Network Name"
    sheet["C1"] = "Product Catagory"
    sheet["D1"] = "Current Version"
    sheet["E1"] = "Latest Stable Version"
    sheet["F1"] = "Scheduled Update"
    line = 2
    #
    for network in results:
        if "network_firmware_check" in results[network].keys():
            for product in results[network]["network_firmware_check"]:
                if product == "is_ok":  # skipping the is_ok key
                    continue
                sheet[f"A{line}"] = org_name
                sheet[f"B{line}"] = network
                sheet[f"C{line}"] = product
                sheet[f"D{line}"] = results[network]["network_firmware_check"][product][
                    "current_version"
                ]
                sheet[f"E{line}"] = results[network]["network_firmware_check"][product][
                    "latest_stable_version"
                ]
                sheet[f"F{line}"] = results[network]["network_firmware_check"][product][
                    "scheduled_upgrade"
                ]
                if (
                    results[network]["network_firmware_check"][product][
                        "current_version"
                    ]
                    != results[network]["network_firmware_check"][product][
                        "latest_stable_version"
                    ]
                ):
                    for cell in sheet[line:line]:
                        cell.font = Font(bold=True, color="00FF9900")
                line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A1:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Network Name"])
    #
    # Channel Utilization tab
    workbook.create_sheet("Channel Utilization")
    sheet = workbook["Channel Utilization"]
    sheet["A1"] = "Organization Name"
    sheet["B1"] = "Network Name"
    sheet["C1"] = "AP Name"
    sheet["D1"] = "Result"
    sheet["E1"] = "Max Utilization"
    sheet["F1"] = "Occurances"
    line = 2
    #
    for network in results:
        if "channel_utilization_check" in results[network].keys():
            for ap in results[network]["channel_utilization_check"]:
                if ap == "is_ok":  # skipping the is_ok key
                    continue
                sheet[f"A{line}"] = org_name
                sheet[f"B{line}"] = network
                sheet[f"C{line}"] = results[network]["channel_utilization_check"][ap][
                    "name"
                ]
                if (
                    results[network]["channel_utilization_check"][ap]["occurances"]
                    < thresholds["5G Occurances Warning"]
                ):
                    sheet[f"D{line}"] = "Pass"
                elif (
                    results[network]["channel_utilization_check"][ap]["occurances"]
                    < thresholds["5G Occurances Alarm"]
                ):
                    sheet[f"D{line}"] = "Fail"
                    sheet[f"D{line}"].font = Font(bold=True, color="00FF6600")
                    sheet[f"E{line}"].font = Font(bold=True, color="00FF6600")
                else:
                    sheet[f"D{line}"] = "Fail"
                    sheet[f"D{line}"].font = Font(bold=True, color="00FF0000")
                    sheet[f"E{line}"].font = Font(bold=True, color="00FF0000")
                sheet[f"E{line}"] = results[network]["channel_utilization_check"][ap][
                    "utilization"
                ]
                sheet[f"F{line}"] = results[network]["channel_utilization_check"][ap][
                    "occurances"
                ]
                if "speed" in results[network]["channel_utilization_check"][ap].keys():
                    sheet["G1"] = "Speed"
                    sheet["H1"] = "Duplex"
                    speed = results[network]["channel_utilization_check"][ap]["speed"]
                    sheet[f"G{line}"] = speed
                    if speed == None:
                        sheet[f"G{line}"].font = Font(bold=True, color="00FF0000")
                    elif speed < 1000:
                        sheet[f"G{line}"].font = Font(bold=True, color="00FF0000")
                    duplex = results[network]["channel_utilization_check"][ap]["duplex"]
                    sheet[f"H{line}"] = duplex
                    if duplex != "full":
                        sheet[f"H{line}"].font = Font(bold=True, color="00FF0000")
                line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A1:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Max Utilization"])
    #
    # RF Profile tab
    workbook.create_sheet("RF Profiles")
    sheet = workbook["RF Profiles"]
    sheet["A1"] = "Organization Name"
    sheet["B1"] = "Network Name"
    sheet["C1"] = "RF Profile"
    sheet["D1"] = "Result"
    sheet["E1"] = "Minimum TX Power"
    sheet["F1"] = "Minimum Bit Rate"
    sheet["G1"] = "Channel Width"
    sheet["H1"] = "RX-SOP"
    line = 2
    #
    for network in results:
        if "rf_profiles_check" in results[network].keys():
            for profile in results[network]["rf_profiles_check"]:
                if profile == "is_ok":  # skipping the is_ok key
                    continue
                sheet[f"A{line}"] = org_name
                sheet[f"B{line}"] = network
                sheet[f"C{line}"] = profile
                if results[network]["rf_profiles_check"][profile]["is_ok"]:
                    sheet[f"D{line}"] = "Pass"
                else:
                    sheet[f"D{line}"] = "Fail"
                    sheet[f"D{line}"].font = Font(bold=True, color="00FF0000")
                sheet[f"E{line}"] = results[network]["rf_profiles_check"][profile][
                    "tests"
                ]["min_power"]["value"]
                if (
                    results[network]["rf_profiles_check"][profile]["tests"][
                        "min_power"
                    ]["is_ok"]
                    == False
                ):
                    sheet[f"E{line}"].font = Font(bold=True, color="00FF0000")
                #
                sheet[f"F{line}"] = results[network]["rf_profiles_check"][profile][
                    "tests"
                ]["min_bitrate"]["value"]
                if (
                    results[network]["rf_profiles_check"][profile]["tests"][
                        "min_bitrate"
                    ]["is_ok"]
                    == False
                ):
                    sheet[f"F{line}"].font = Font(bold=True, color="00FF0000")
                #
                sheet[f"G{line}"] = results[network]["rf_profiles_check"][profile][
                    "tests"
                ]["channel_width"]["value"]
                if (
                    results[network]["rf_profiles_check"][profile]["tests"][
                        "channel_width"
                    ]["is_ok"]
                    == False
                ):
                    sheet[f"G{line}"].font = Font(bold=True, color="00FF0000")
                #
                sheet[f"H{line}"] = results[network]["rf_profiles_check"][profile][
                    "tests"
                ]["rxsop"]["value"]
                if (
                    results[network]["rf_profiles_check"][profile]["tests"]["rxsop"][
                        "is_ok"
                    ]
                    == False
                ):
                    sheet[f"H{line}"].font = Font(bold=True, color="00FF0000")
                line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A1:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Result"])
    #
    # Switch ports tab
    workbook.create_sheet("Switch port counters")
    sheet = workbook["Switch port counters"]
    sheet["A1"] = "Organization Name"
    sheet["B1"] = "Network Name"
    sheet["C1"] = "Switch"
    sheet["D1"] = "Result"
    sheet["E1"] = "Ports with CRC errors"
    sheet["F1"] = "Ports with collisions"
    sheet["G1"] = "Multicasts exceeding threshold"
    sheet["H1"] = "Broadcasts exceeding threshold"
    sheet["I1"] = "Topology changes exceeding threshold"
    line = 2
    #
    for network in results:
        if "port_counters_check" in results[network].keys():
            for switch in results[network]["port_counters_check"]:
                if switch == "is_ok":  # skipping the is_ok key
                    continue
                sheet[f"A{line}"] = org_name
                sheet[f"B{line}"] = network
                sheet[f"C{line}"] = switch
                if results[network]["port_counters_check"][switch]["is_ok"]:
                    sheet[f"D{line}"] = "Pass"
                else:
                    sheet[f"D{line}"] = "Fail"
                    sheet[f"D{line}"].font = Font(bold=True, color="00FF0000")
                if results[network]["port_counters_check"][switch]["crc"] != []:
                    sheet[f"E{line}"] = str(
                        results[network]["port_counters_check"][switch]["crc"]
                    )
                    sheet[f"E{line}"].font = Font(bold=True, color="00FF0000")
                if results[network]["port_counters_check"][switch]["collision"] != []:
                    sheet[f"F{line}"] = str(
                        results[network]["port_counters_check"][switch]["collision"]
                    )
                    sheet[f"F{line}"].font = Font(bold=True, color="00FF0000")
                if results[network]["port_counters_check"][switch]["multicast"] != []:
                    sheet[f"G{line}"] = str(
                        results[network]["port_counters_check"][switch]["multicast"]
                    )
                    sheet[f"G{line}"].font = Font(bold=True, color="00FF0000")
                if results[network]["port_counters_check"][switch]["broadcast"] != []:
                    sheet[f"H{line}"] = str(
                        results[network]["port_counters_check"][switch]["broadcast"]
                    )
                    sheet[f"H{line}"].font = Font(bold=True, color="00FF0000")
                if (
                    results[network]["port_counters_check"][switch]["topology_changes"]
                    != []
                ):
                    sheet[f"I{line}"] = str(
                        results[network]["port_counters_check"][switch][
                            "topology_changes"
                        ]
                    )
                    sheet[f"I{line}"].font = Font(bold=True, color="00FF0000")
                line += 1
    #
    # Adding filters
    sheet.auto_filter.ref = f"A1:{ABC[sheet.max_column]}{line}"
    sheet.auto_filter.add_filter_column(0, ["Result"])
    #
    # Formatting: increasing font size, adjusting column width
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                cell.font = Font(size=16, bold=cell.font.bold, color=cell.font.color)
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            if not "Introduction" in sheet_name:
                sheet.column_dimensions[column[0].column_letter].width = max_length + 10
    #
    sheet = workbook["Introduction"]
    sheet["B3"].font = Font(bold=True, size=36)
    #
    workbook.save(filename=f"{org_name}.xlsx")


async def main():
    async with meraki.aio.AsyncDashboardAPI(
        output_log=False,
        suppress_logging=True,
        maximum_concurrent_requests=5,
        wait_on_rate_limit=True,
        nginx_429_retry_wait_time=2,
        maximum_retries=100,
        caller='Healthcheck/22.12 OBrigg'
    ) as aiomeraki:
        #
        # Checking the organization has API enabled
        try:
            dashboard.organizations.getOrganizationNetworks(org_id)
        except meraki.APIError as e:
            pp(f"[red]An error has occured: \n\n{e}[/red]\n\n")
            sys.exit(1)
        #
        # Run organization checks
        await async_check_org_admins(aiomeraki)
        # Get networks
        try:
            networks = await aiomeraki.organizations.getOrganizationNetworks(org_id)
        except meraki.exceptions.AsyncAPIError as e:
            pp(
                f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
            )
        except Exception as e:
            pp(f"[bold magenta]Some other ERROR: {e}")

        # Prepare results per network
        for network in networks:
            results[network["name"]] = {}
        #
        # General checks
        check_network_health_tasks = [
            async_check_network_health_alerts(aiomeraki, net) for net in networks
        ]
        for task in asyncio.as_completed(check_network_health_tasks):
            await task
        check_firmware_tasks = [
            async_check_network_firmware(aiomeraki, net) for net in networks
        ]
        for task in asyncio.as_completed(check_firmware_tasks):
            await task
        #
        # Wired checks
        check_switch_port_tasks = [
            async_check_switches_port_counters(aiomeraki, net)
            for net in networks
            if "switch" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_switch_port_tasks):
            await task
        check_switch_stp_tasks = [
            async_check_switch_stp(aiomeraki, net)
            for net in networks
            if "switch" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_switch_stp_tasks):
            await task
        check_switch_mtu_tasks = [
            async_check_switch_mtu(aiomeraki, net)
            for net in networks
            if "switch" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_switch_mtu_tasks):
            await task
        check_switch_storm_control_tasks = [
            async_check_switch_storm_control(aiomeraki, net)
            for net in networks
            if "switch" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_switch_storm_control_tasks):
            await task
        #
        # Wireless checks
        check_channel_utilization_tasks = [
            async_check_wifi_channel_utilization(aiomeraki, net)
            for net in networks
            if "wireless" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_channel_utilization_tasks):
            await task
        check_rf_profiles_tasks = [
            async_check_wifi_rf_profiles(aiomeraki, net)
            for net in networks
            if "wireless" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_rf_profiles_tasks):
            await task
        check_ssid_amount_tasks = [
            async_check_wifi_ssid_amount(aiomeraki, net)
            for net in networks
            if "wireless" in net["productTypes"]
        ]
        for task in asyncio.as_completed(check_ssid_amount_tasks):
            await task
        # TODO: wireless health
        # 
        # Early API Access
        is_early_api_access_enabled = False
        try:
            early_access = dashboard.organizations.getOrganizationEarlyAccessFeatures(org_id)
            for feature in early_access:
                if feature['shortName'] == 'has_beta_api' and 'isOrgScopedOnly':
                    is_early_api_access_enabled = True
            #            
            if is_early_api_access_enabled:
                headers = {'Content-Type': 'application/json', 
                            'Accept': 'application/json', 
                            'X-Cisco-Meraki-API-Key': api_key}
                check_wireless_ports(headers)
        except meraki.exceptions.AsyncAPIError as e:
            pp(
                f'[bold magenta]Meraki AIO API Error (OrgID "{ org_id }", OrgName "{ org_name }"): \n { e }'
            )
        except Exception as e:
            pp(f"[bold magenta]Some other ERROR: {e}")
        #
        pp("\n", 100 * "*", "\n")
        # Results cleanup
        clean_results = {}
        for result in results:
            if results[result] != {}:
                clean_results[result] = results[result]

        pp(clean_results)
        generate_excel_report(clean_results)
        pp("Done.")


if __name__ == "__main__":
    # Thresholds
    thresholds = {
        "5G Channel Utilization": 20,  # %
        "5G Occurances Warning": 10,  # times
        "5G Occurances Alarm": 50,  # times
        "5G Min TX Power": 10,  # dBm
        "5G Min Bitrate": 12,  # Mbps
        "5G Max Channel Width": 40,  # MHz
        "broadcast_rate": 100,  # pps
        "multicast_rate": 100,  # pps
        "topology_changes": 10,
        "ssid_amount": 4,
    }
    results = {}

    # Check for an envriomnet variable, if not set, ask for an API key
    if os.environ.get("MERAKI_DASHBOARD_API_KEY"):
        api_key = os.environ["MERAKI_DASHBOARD_API_KEY"]
    else:
        pp(
            "[bold magenta]No API key found. Please enter your Meraki Dashboard API key:"
        )
        api_key = getpass("Meraki Dashboard API Key: ")
        os.environ["MERAKI_DASHBOARD_API_KEY"] = api_key
    # Initializing Meraki SDK
    dashboard = meraki.DashboardAPI(output_log=False, suppress_logging=True, caller='Healthcheck/22.12 OBrigg')
    org_id, org_name = select_org()

    start = time.time()
    loop = asyncio.new_event_loop()
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())

    end = time.time()
    pp(f"Total run time: {round(end - start, 2)} seconds")
